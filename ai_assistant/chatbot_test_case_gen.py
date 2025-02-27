import os
import pypdf
import streamlit as st
import pandas as pd
import docx
from io import StringIO, BytesIO
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook

# Set up OpenAI API key
os.environ["OPENAI_API_KEY"] = st.secrets["OPENAI_API_KEY"]

# Define the chat prompt template
st.logo("./bot.png")
st.title("Generate Your Manual Test Cases...")
st.caption("Bot can make mistakes. Review the response before using.")

# Initialize the model
llm = ChatOpenAI(model="gpt-3.5-turbo")

# Initialize session state
if "conversation_history" not in st.session_state:
    st.session_state.conversation_history = []
if "test_cases" not in st.session_state:
    st.session_state.test_cases = []  # Store structured test cases


# Function to extract text from files
def extract_text_from_file(uploaded_file):
    file_type = uploaded_file.name.split(".")[-1].lower()

    if file_type == "docx":
        doc = docx.Document(uploaded_file)
        return "\n".join([para.text for para in doc.paragraphs]).strip()

    elif file_type == "pdf":
        pdf_reader = pypdf.PdfReader(uploaded_file)
        return "\n".join([page.extract_text() for page in pdf_reader.pages if page.extract_text()]).strip()

    elif file_type == "txt":
        return StringIO(uploaded_file.getvalue().decode("utf-8")).read().strip()

    elif file_type == "csv":
        csv_data = pd.read_csv(uploaded_file)
        return csv_data.to_string(index=False).strip()

    elif file_type == "xlsx":
        wb = load_workbook(uploaded_file, data_only=True)
        sheet = wb.active
        return "\n".join(["\t".join([str(cell.value) for cell in row]) for row in sheet.iter_rows()]).strip()

    else:
        return "Unsupported file format!"


# Function to parse test cases from bot response
def parse_test_cases(response_text):
    test_cases = []
    test_case_blocks = response_text.strip().split("\n\n")  # Splitting based on blank lines

    for block in test_case_blocks:
        lines = block.split("\n")
        test_case = {}
        key = None

        for line in lines:
            if ":" in line:  # Identify key-value pairs
                parts = line.split(":", 1)
                key, value = parts[0].strip(), parts[1].strip()
                test_case[key] = value
            elif key:  # Handle multi-line values (e.g., Test Steps, Test Data)
                test_case[key] += f"\n{line.strip()}"

        # Ensure all required fields are captured
        required_fields = ["Test Case ID", "Test Case Title", "Pre-conditions", "Test Data", "Test Steps",
                           "Expected Result", "Priority"]
        if all(k in test_case for k in required_fields):
            test_cases.append(test_case)

    return test_cases


# Function to handle input
def handle_input(input_text, uploaded_file):
    if uploaded_file:
        input_text = extract_text_from_file(uploaded_file)

    if input_text:
        st.session_state.conversation_history.append(("user", input_text))
        actual_requirement = input_text

        test_case_format = """
        Test Case ID: (Unique identifier, e.g., TC_01)
        Test Case Title: (Briefly describe what the test case validates)
        Pre-conditions: (Any setup or pre-requirements before executing the test case)
        Test Data: (Data used in the test case)
        Test Steps: (Step-by-step actions to execute the test case)
        Expected Result: (The expected outcome after executing the steps)
        Priority: (Low/Medium/High based on impact)
        """

        if not input_text.lower().startswith("generate test cases"):
            input_text = (f"You are a manual software tester. Please review the given requirements carefully and write "
                          f"detailed manual test cases for {input_text} in the format {test_case_format}. Ensure "
                          f"comprehensive coverage, including positive, negative, boundary, and edge cases where "
                          f"applicable.")

        chat_history = [
                           ("system",
                            f"You are a manual software tester. Please review the given requirements carefully and "
                            f"write detailed manual test cases for {input_text} in the format {test_case_format}. "
                            f"Ensure comprehensive coverage, including positive, negative, boundary, and edge cases "
                            f"where applicable."),
                       ] + [(role, msg) for role, msg in st.session_state.conversation_history if role == "user"]

        try:
            modified_prompt = ChatPromptTemplate.from_messages(chat_history)
            modified_chain = modified_prompt | llm
            response = modified_chain.invoke({"question": input_text})
            content = getattr(response, "content", None) or response.get("content", None)

            if content:
                formatted_content = f"Generated test cases for **{actual_requirement}**\n\n{content}"
                st.session_state.conversation_history.append(("assistant", formatted_content))

                # Parse and store test cases
                test_cases = parse_test_cases(content)
                st.session_state.test_cases.extend(test_cases)

            else:
                st.session_state.conversation_history.append(("assistant", "No valid response received."))
        except Exception as e:
            st.session_state.conversation_history.append(("assistant", f"Error: {e}"))


# User Input & File Upload
user_input = st.chat_input("Describe requirement here")
uploaded_file = st.file_uploader("Upload the requirement file", type=["docx", "pdf", "txt", "csv", "xlsx"])

# Process input or file
if user_input or uploaded_file:
    handle_input(user_input, uploaded_file)

# Display conversation history
if st.session_state.conversation_history:
    for role, message in st.session_state.conversation_history:
        with st.chat_message(role):
            st.markdown(message)

# Export test cases to Excel
if st.session_state.test_cases:
    df = pd.DataFrame(st.session_state.test_cases)

    # Save DataFrame to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Test Cases")

    st.download_button(
        label="Download Test Cases",
        data=output.getvalue(),
        file_name="test_cases.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
