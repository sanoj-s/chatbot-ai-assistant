import os
from datetime import datetime
from deepeval import assert_test
from deepeval.metrics import FaithfulnessMetric, AnswerRelevancyMetric, SummarizationMetric, HallucinationMetric, \
    ContextualRelevancyMetric, ContextualRecallMetric, ContextualPrecisionMetric, BiasMetric, ToxicityMetric, GEval
from deepeval.metrics.ragas import RAGASAnswerRelevancyMetric, RAGASFaithfulnessMetric, \
    RAGASContextualRecallMetric, RAGASContextualPrecisionMetric
from deepeval.test_case import LLMTestCase, LLMTestCaseParams
from openpyxl.workbook import Workbook
from datasets import Dataset
from ragas import evaluate
from ragas.metrics import faithfulness, answer_correctness, answer_similarity, context_precision, answer_relevancy, \
    context_recall, context_entity_recall


def evaluate_faithfulness_deepeval(user_input_data_list, retrieval_context_data_list, bot_response_data_list,
                                   threshold_value=0.5,
                                   model_value="gpt-4"):
    """
    Evaluate the faithfulness of the bot's response to the user's input.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param retrieval_context_data_list:
    :param user_input_data_list: List of user inputs
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for relevancy
    :param model_value: Model to evaluate relevancy
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_faithfulness = FaithfulnessMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"FaithfulnessReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Retrieval Context", "Actual Output", "Faithfulness Score", "Reason"])

    # Evaluate each user input and bot response
    for user_input_data, retrieval_context, bot_response_data in zip(user_input_data_list, retrieval_context_data_list,
                                                                     bot_response_data_list):
        # If retrieval_context is a list, convert it to a string
        context_str = ', '.join(retrieval_context) if isinstance(retrieval_context, list) else retrieval_context

        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data,
                                retrieval_context=retrieval_context)
        # Measure the metrics
        metric_faithfulness.measure(test_case)

        # Extract the score
        score = metric_faithfulness.score
        reason = metric_faithfulness.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_faithfulness])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_answer_relevancy_deepeval(user_input_data_list, bot_response_data_list, threshold_value=0.5,
                                       model_value="gpt-4"):
    """
    Evaluate the relevancy of the bot's responses to the user's inputs and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for relevancy
    :param model_value: Model to evaluate relevancy
    :return: None
    """
    # Validate input lengths
    if len(user_input_data_list) != len(bot_response_data_list):
        raise ValueError("The number of user inputs and bot responses must be the same.")

    # Initialize the metric
    metric_answer_relevancy = AnswerRelevancyMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"AnswerRelevancyReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Actual Output", "Answer Relevancy Score", "Reason"])

    # Evaluate each user input and bot response
    for user_input_data, bot_response_data in zip(user_input_data_list, bot_response_data_list):
        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data)

        # Measure the answer relevancy
        metric_answer_relevancy.measure(test_case)

        # Extract the score
        score = metric_answer_relevancy.score
        reason = metric_answer_relevancy.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, bot_response_data, score, reason])

        # Assert test for relevancy check
        try:
            assert_test(test_case, [metric_answer_relevancy])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_summarization_deepeval(user_input_data_list, assessment_questions_data_list, bot_response_data_list,
                                    threshold_value=0.7, model_value="gpt-4"):
    """
    Evaluate the summarization of the bot's response to the user's input.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param assessment_questions_data_list: List of assessment questions
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate summarization
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(bot_response_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_summarization = SummarizationMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"SummarizationReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Assessment Questions", "Bot Response", "Summarization Score", "Reason"])

    # Evaluate each user input and bot response
    for user_input_data, assessment_questions_data, bot_response_data in zip(
            user_input_data_list, assessment_questions_data_list, bot_response_data_list):
        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data)

        # Dynamic assessment questions for summarization
        metric_summarization.assessment_questions = [assessment_questions_data]

        # Measure the summarization quality
        metric_summarization.measure(test_case)

        # Extract the score
        score = metric_summarization.score
        reason = metric_summarization.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, assessment_questions_data, bot_response_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_summarization])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_hallucination_deepeval(user_input_data_list, context_data_list, bot_response_data_list,
                                    threshold_value=0.7, model_value="gpt-4"):
    """
    Evaluate the hallucination of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param context_data_list: List of contexts
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate hallucination
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(context_data_list) == len(bot_response_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_hallucination = HallucinationMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"HallucinationReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Context", "Bot Response", "Hallucination Score", "Reason"])

    # Evaluate each test case
    for user_input_data, context_data, bot_response_data in zip(user_input_data_list, context_data_list,
                                                                bot_response_data_list):

        # If retrieval_context is a list, convert it to a string
        context_str = ', '.join(context_data) if isinstance(context_data, list) else context_data

        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data, context=context_data)

        # Measure hallucination
        metric_hallucination.measure(test_case)

        # Extract the score and reason
        score = metric_hallucination.score
        reason = metric_hallucination.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_hallucination])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_contextual_relevancy_deepeval(user_input_data_list, retrieval_context_data_list, bot_response_data_list,
                                           threshold_value=0.7, model_value="gpt-4"):
    """
    Evaluate the contextual relevancy of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate contextual relevancy
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_contextual_relevancy = ContextualRelevancyMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ContextualRelevancyReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Retrieval Context", "Bot Response", "Contextual Relevancy Score", "Reason"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list):

        # If retrieval_context is a list, convert it to a string
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data

        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data,
                                retrieval_context=retrieval_context_data)

        # Measure contextual relevancy
        metric_contextual_relevancy.measure(test_case)

        # Extract the score and reason
        score = metric_contextual_relevancy.score
        reason = metric_contextual_relevancy.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_contextual_relevancy])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_contextual_recall_deepeval(user_input_data_list, retrieval_context_data_list, bot_response_data_list,
                                        expected_output_data_list, threshold_value=0.7, model_value="gpt-4"):
    """
    Evaluate the contextual recall of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param expected_output_data_list: List of expected outputs
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate contextual recall
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_contextual_recall = ContextualRecallMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ContextualRecallReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Retrieval Context", "Bot Response", "Expected Output", "Recall Score", "Reason"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list, expected_output_data_list):

        # If retrieval_context is a list, convert it to a string
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data
        # Create the test case
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data,
            retrieval_context=retrieval_context_data
        )

        # Measure Contextual Recall
        metric_contextual_recall.measure(test_case)

        # Extract the score and reason
        score = metric_contextual_recall.score
        reason = metric_contextual_recall.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, expected_output_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_contextual_recall])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_contextual_precision_deepeval(user_input_data_list, retrieval_context_data_list, bot_response_data_list,
                                           expected_output_data_list, threshold_value=0.7, model_value="gpt-4"):
    """
    Evaluate the contextual precision of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param expected_output_data_list: List of expected outputs
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate contextual precision
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_contextual_precision = ContextualPrecisionMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ContextualPrecisionReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Retrieval Context", "Bot Response", "Expected Output", "Precision Score", "Reason"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list, expected_output_data_list):

        # If retrieval_context is a list, convert it to a string
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data

        # Create the test case
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data,
            retrieval_context=retrieval_context_data
        )

        # Measure Contextual Precision
        metric_contextual_precision.measure(test_case)

        # Extract the score and reason
        score = metric_contextual_precision.score
        reason = metric_contextual_precision.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, expected_output_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_contextual_precision])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_bias_deepeval(user_input_data_list, bot_response_data_list, threshold_value=0.5, model_value="gpt-4"):
    """
    Evaluate the bias of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate bias
    :return: None
    """
    # Validate input lengths
    if len(user_input_data_list) != len(bot_response_data_list):
        raise ValueError("User input data and bot response data must have the same length.")

    # Initialize the metric
    metric_bias = BiasMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"BiasEvaluationReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Bot Response", "Bias Score", "Reason"])

    # Evaluate each test case
    for user_input_data, bot_response_data in zip(user_input_data_list, bot_response_data_list):
        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data)

        # Measure Bias
        metric_bias.measure(test_case)

        # Extract the score and reason
        score = metric_bias.score
        reason = metric_bias.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, bot_response_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_bias])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_toxicity_deepeval(user_input_data_list, bot_response_data_list, threshold_value=0.5, model_value="gpt-4"):
    """
    Evaluate the toxicity of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param bot_response_data_list: List of bot responses
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate toxicity
    :return: None
    """
    # Validate input lengths
    if len(user_input_data_list) != len(bot_response_data_list):
        raise ValueError("User input data and bot response data must have the same length.")

    # Initialize the metric
    metric_toxicity = ToxicityMetric(
        threshold=threshold_value,
        model=model_value,
        include_reason=True
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ToxicityEvaluationReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Bot Response", "Toxicity Score", "Reason"])

    # Evaluate each test case
    for user_input_data, bot_response_data in zip(user_input_data_list, bot_response_data_list):
        # Create the test case
        test_case = LLMTestCase(input=user_input_data, actual_output=bot_response_data)

        # Measure Toxicity
        metric_toxicity.measure(test_case)

        # Extract the score and reason
        score = metric_toxicity.score
        reason = metric_toxicity.reason

        # Append results to Excel sheet
        sheet.append([user_input_data, bot_response_data, score, reason])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_toxicity])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_ragas_answer_relevancy_metrics_deepeval(user_input_data_list, retrieval_context_data_list,
                                                     bot_response_data_list,
                                                     expected_output_data_list, threshold_value=0.5,
                                                     model_value="gpt-4"):
    """
    Evaluate the Ragas Answer Relevancy metrics of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param expected_output_data_list: List of expected outputs
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate relevancy
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_ragas = RAGASAnswerRelevancyMetric(
        threshold=threshold_value,
        model=model_value
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"RAGASAnswerRelevancyReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Retrieval Context", "Bot Response", "Expected Output", "Relevancy Score"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list, expected_output_data_list):

        # Convert retrieval_context_data to string if it's a list
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data

        # Create the test case
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data,
            retrieval_context=retrieval_context_data
        )

        # Measure Ragas Answer Relevancy metrics
        metric_ragas.measure(test_case)

        # Extract the score
        score = metric_ragas.score

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, expected_output_data, score])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_ragas])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_ragas_faithfulness_metrics_deepeval(user_input_data_list, retrieval_context_data_list,
                                                 bot_response_data_list,
                                                 expected_output_data_list, threshold_value=0.5, model_value="gpt-4"):
    """
    Evaluate the Ragas Faithfulness metrics of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param expected_output_data_list: List of expected outputs
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate faithfulness
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_ragas = RAGASFaithfulnessMetric(
        threshold=threshold_value,
        model=model_value
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"RAGASFaithfulnessReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Retrieval Context", "Bot Response", "Expected Output", "Faithfulness Score"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list, expected_output_data_list):

        # Convert retrieval_context_data to string if it's a list
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data

        # Create the test case
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data,
            retrieval_context=retrieval_context_data
        )

        # Measure Ragas Faithfulness metrics
        metric_ragas.measure(test_case)

        # Extract the score
        score = metric_ragas.score

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, expected_output_data, score])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_ragas])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_ragas_contextual_recall_metrics_deepeval(user_input_data_list, retrieval_context_data_list,
                                                      bot_response_data_list,
                                                      expected_output_data_list, threshold_value=0.5,
                                                      model_value="gpt-4"):
    """
    Evaluate the Ragas Contextual Recall metrics of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param expected_output_data_list: List of expected outputs
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate contextual recall
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_ragas = RAGASContextualRecallMetric(
        threshold=threshold_value,
        model=model_value
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"RAGASContextualRecallReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(
        ["User Input", "Retrieval Context", "Bot Response", "Expected Output", "Contextual Recall Score"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list, expected_output_data_list):

        # Convert retrieval_context_data to string if it's a list
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data

        # Create the test case
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data,
            retrieval_context=retrieval_context_data
        )

        # Measure Ragas Contextual Recall metrics
        metric_ragas.measure(test_case)

        # Extract the score
        score = metric_ragas.score

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, expected_output_data, score])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_ragas])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_ragas_contextual_precision_metrics_deepeval(user_input_data_list, retrieval_context_data_list,
                                                         bot_response_data_list,
                                                         expected_output_data_list, threshold_value=0.5,
                                                         model_value="gpt-4"):
    """
    Evaluate the Ragas Contextual Precision metrics of the bot's response to the user's input and generate a report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list: List of user inputs
    :param retrieval_context_data_list: List of retrieval contexts
    :param bot_response_data_list: List of bot responses
    :param expected_output_data_list: List of expected outputs
    :param threshold_value: Threshold for evaluation
    :param model_value: Model to evaluate contextual precision
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    # Initialize the metric
    metric_ragas = RAGASContextualPrecisionMetric(
        threshold=threshold_value,
        model=model_value
    )

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"RAGASContextualPrecisionReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(
        ["User Input", "Retrieval Context", "Bot Response", "Expected Output", "Contextual Precision Score"])

    # Evaluate each test case
    for user_input_data, retrieval_context_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, retrieval_context_data_list, bot_response_data_list, expected_output_data_list):

        # Convert retrieval_context_data to string if it's a list
        context_str = ', '.join(retrieval_context_data) if isinstance(retrieval_context_data,
                                                                      list) else retrieval_context_data

        # Create the test case
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data,
            retrieval_context=retrieval_context_data
        )

        # Measure Ragas Contextual Precision metrics
        metric_ragas.measure(test_case)

        # Extract the score
        score = metric_ragas.score

        # Append results to Excel sheet
        sheet.append([user_input_data, context_str, bot_response_data, expected_output_data, score])

        # Assert test for metrics
        try:
            assert_test(test_case, [metric_ragas])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_metrics_using_g_eval_deepeval(metric_name, criteria_details, user_input_data_list, bot_response_data_list,
                                           expected_output_data_list):
    """
    Evaluate the metrics using G-Eval and generate an Excel report.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param metric_name: Name of the metric being evaluated.
    :param criteria_details: The criteria details used for evaluation.
    :param user_input_data_list: List of user input data for multiple test cases.
    :param bot_response_data_list: List of bot responses for multiple test cases.
    :param expected_output_data_list: List of expected output data for multiple test cases.
    :return: None
    """
    # Validate input lengths
    if not (len(user_input_data_list) == len(bot_response_data_list) == len(expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"G-Eval_Metrics_Report_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Write header
    sheet.append(["User Input", "Bot Response", "Expected Output", "Metric Score", "Reason"])

    # Evaluate each test case
    for user_input_data, bot_response_data, expected_output_data in zip(
            user_input_data_list, bot_response_data_list, expected_output_data_list):

        # Initialize the metric
        correctness_metric = GEval(
            name=metric_name,
            criteria=criteria_details,
            evaluation_params=[LLMTestCaseParams.INPUT, LLMTestCaseParams.ACTUAL_OUTPUT,
                               LLMTestCaseParams.EXPECTED_OUTPUT]
        )

        # Create the test case for metric evaluation using G-Eval
        test_case = LLMTestCase(
            input=user_input_data,
            actual_output=bot_response_data,
            expected_output=expected_output_data
        )

        # Measure the metric using G-Eval
        correctness_metric.measure(test_case)

        # Extract the score and reason
        score = correctness_metric.score
        reason = correctness_metric.reason

        # Append the results to the Excel sheet
        sheet.append([user_input_data, bot_response_data, expected_output_data, score, reason])

        # Print the score and reason for the metric
        print(f"Score: {score}, Reason: {reason}")

        # Assert test for metrics
        try:
            assert_test(test_case, [correctness_metric])
        except AssertionError:
            print(f"Assertion failed for input: {user_input_data}\nBot response: {bot_response_data}")
            continue

    # Save the Excel file
    workbook.save(file_path)
    print(f"Report generated: {file_path}")


def evaluate_all_metrics_deepeval(user_input_data_list, retrieval_context_data_list, bot_response_data_list,
                                  expected_output_data_list, metric_name, criteria_details,
                                  threshold_value=0.7, model_value="gpt-4"):
    """
    Evaluate all metrics based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param user_input_data_list:
    :param retrieval_context_data_list:
    :param bot_response_data_list:
    :param expected_output_data_list:
    :param metric_name:
    :param criteria_details:
    :param threshold_value:
    :param model_value:
    :return:
    """
    # Initialize all metrics
    metrics = [
        FaithfulnessMetric(threshold=threshold_value, model=model_value, include_reason=True),
        AnswerRelevancyMetric(threshold=threshold_value, model=model_value, include_reason=True),
        HallucinationMetric(threshold=threshold_value, model=model_value, include_reason=True),
        ContextualRelevancyMetric(threshold=threshold_value, model=model_value, include_reason=True),
        ContextualRecallMetric(threshold=threshold_value, model=model_value, include_reason=True),
        ContextualPrecisionMetric(threshold=threshold_value, model=model_value, include_reason=True),
        BiasMetric(threshold=threshold_value, model=model_value, include_reason=True),
        ToxicityMetric(threshold=threshold_value, model=model_value, include_reason=True),
        RAGASAnswerRelevancyMetric(threshold=threshold_value, model=model_value),
        RAGASFaithfulnessMetric(threshold=threshold_value, model=model_value),
        RAGASContextualRecallMetric(threshold=threshold_value, model=model_value),
        RAGASContextualPrecisionMetric(threshold=threshold_value, model=model_value),
        GEval(name=metric_name, criteria=criteria_details, evaluation_params=[LLMTestCaseParams.INPUT,
                                                                              LLMTestCaseParams.ACTUAL_OUTPUT,
                                                                              LLMTestCaseParams.EXPECTED_OUTPUT])
    ]

    # Validate input lengths
    if not (len(user_input_data_list) == len(retrieval_context_data_list) == len(bot_response_data_list) == len(
            expected_output_data_list)):
        raise ValueError("All input lists must have the same length.")

    report_folder_path = __get_deepeval_report_path()

    # Prepare Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"AllMetricsReport_{timestamp}.xlsx"
    file_path = os.path.join(report_folder_path, file_name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Metrics Results"

    # Write header
    header = ["User Input", "Context", "Bot Response", "Expected Output"]
    for metric in metrics:
        header.append(f"{metric.__class__.__name__} Score")
    sheet.append(header)

    # Iterate over test cases
    for user_input, retrieval_context, bot_response, expected_output in zip(user_input_data_list,
                                                                            retrieval_context_data_list,
                                                                            bot_response_data_list,
                                                                            expected_output_data_list):

        # If retrieval_context is a list, convert it to a string
        context_str = ', '.join(retrieval_context) if isinstance(retrieval_context, list) else retrieval_context

        # Create the test case for the current set of data
        test_case = LLMTestCase(
            input=user_input,
            actual_output=bot_response,
            retrieval_context=retrieval_context,
            context=retrieval_context,
            expected_output=expected_output
        )

        # Collect results for this test case
        results_row = [user_input, context_str, bot_response, expected_output]
        for metric in metrics:
            try:
                metric.measure(test_case)
                results_row.append(metric.score)

                # Assert the metric
                assert_test(test_case, [metric])
            except AssertionError:
                # Handle failed assertions
                print(f"Assertion failed for {metric.__class__.__name__} on test case.")
                continue

        # Append the results row to the sheet
        sheet.append(results_row)

    # Save the Excel file
    workbook.save(file_path)
    print(f"Consolidated Metrics report generated: {file_path}")


def evaluate_all_metrics_ragas(data_samples):
    """
    Evaluate all metrics based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[context_precision, context_recall, context_entity_recall, faithfulness, answer_relevancy,
                              answer_correctness, answer_similarity])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"AllMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_context_precision_ragas(data_samples):
    """
    Evaluate context precision based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[context_precision])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ContextPrecisionMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_context_recall_ragas(data_samples):
    """
    Evaluate context recall based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[context_recall])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ContextRecallMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_context_entity_recall_ragas(data_samples):
    """
    Evaluate context entity recall based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[context_entity_recall])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"ContextEntityRecallMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_faithfulness_ragas(data_samples):
    """
    Evaluate faithfulness based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[faithfulness])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"FaithfulnessMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_answer_relevancy_ragas(data_samples):
    """
    Evaluate answer relevancy based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[answer_relevancy])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"AnswerRelevancyMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_answer_correctness_ragas(data_samples):
    """
    Evaluate answer correctness based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[answer_correctness])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"AnswerCorrectnessMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def evaluate_answer_similarity_ragas(data_samples):
    """
    Evaluate answer similarity based on the given user's input, bot's response, context, and ground truth,
    and generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param data_samples:
    :return:
    """
    dataset = Dataset.from_dict(data_samples)
    score = evaluate(dataset,
                     metrics=[answer_similarity])
    df = score.to_pandas()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"AnswerSimilarityMetricsReport_{timestamp}.xlsx"
    __generate_report(df, file_name)


def __generate_report(df, file_name):
    """
    Generate an Excel report with the results for multiple test cases.
    @Author: Sanoj Swaminathan
    @Date: 06-12-2024
    :param df:
    :param file_name:
    :return:
    """
    report_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reports_ragas')
    os.makedirs(report_dir, exist_ok=True)
    df.to_excel(os.path.join(report_dir, file_name), index=False)


def __get_deepeval_report_path():
    """
    Get the path of the DeepEval report folder
    @Author: Sanoj Swaminathan
    @Date: 09-12-2024
    :return:
    """
    report_folder = os.path.join(os.getcwd(), "../reports_deepeval")
    if not os.path.exists(report_folder):
        os.makedirs(report_folder)
    return report_folder
